import os
import subprocess
from pathlib import Path

# Supported conversions map: input_ext -> [output_formats]
SUPPORTED_CONVERSIONS = {
    # Images
    "jpg":  ["png", "webp", "bmp", "gif", "tiff", "ico", "pdf"],
    "jpeg": ["png", "webp", "bmp", "gif", "tiff", "ico", "pdf"],
    "png":  ["jpg", "webp", "bmp", "gif", "tiff", "ico", "pdf"],
    "webp": ["jpg", "png", "bmp", "gif", "tiff"],
    "bmp":  ["jpg", "png", "webp", "gif", "tiff", "ico"],
    "gif":  ["jpg", "png", "webp", "mp4"],
    "tiff": ["jpg", "png", "webp", "bmp", "pdf"],
    "tif":  ["jpg", "png", "webp", "bmp", "pdf"],
    "ico":  ["jpg", "png", "webp", "bmp"],
    "svg":  ["jpg", "png", "pdf"],
    "heic": ["jpg", "png"],
    "heif": ["jpg", "png"],

    # Documents
    "pdf":  ["docx", "txt", "jpg", "png"],
    "docx": ["pdf", "txt", "html"],
    "doc":  ["pdf", "txt", "html"],
    "txt":  ["pdf", "docx", "html"],
    "html": ["pdf", "txt"],
    "htm":  ["pdf", "txt"],
    "xlsx": ["csv", "pdf"],
    "xls":  ["csv", "pdf"],
    "csv":  ["xlsx", "pdf"],
    "pptx": ["pdf"],
    "ppt":  ["pdf"],
    "odt":  ["pdf", "docx", "txt"],
    "rtf":  ["pdf", "txt", "docx"],

    # Audio
    "mp3":  ["wav", "ogg", "flac", "aac", "m4a", "opus"],
    "wav":  ["mp3", "ogg", "flac", "aac", "m4a", "opus"],
    "ogg":  ["mp3", "wav", "flac", "aac"],
    "flac": ["mp3", "wav", "ogg", "aac", "m4a"],
    "aac":  ["mp3", "wav", "ogg", "flac"],
    "m4a":  ["mp3", "wav", "ogg", "flac", "aac"],
    "opus": ["mp3", "wav", "ogg"],
    "wma":  ["mp3", "wav", "ogg", "flac"],

    # Video
    "mp4":  ["avi", "mov", "mkv", "webm", "gif", "mp3", "wav"],
    "avi":  ["mp4", "mov", "mkv", "webm", "mp3"],
    "mov":  ["mp4", "avi", "mkv", "webm", "mp3"],
    "mkv":  ["mp4", "avi", "mov", "webm", "mp3"],
    "webm": ["mp4", "avi", "mov", "mkv", "gif"],
    "flv":  ["mp4", "avi", "mkv", "mp3"],
    "wmv":  ["mp4", "avi", "mkv", "mp3"],
    "3gp":  ["mp4", "avi", "mp3"],
}


def get_output_formats(input_ext: str) -> list:
    """Get available output formats for a given input extension."""
    return SUPPORTED_CONVERSIONS.get(input_ext.lower(), [])


def convert_file(input_path: str, input_ext: str, output_path: str, output_fmt: str) -> tuple:
    """
    Convert a file from input format to output format.
    Returns (success: bool, error_message: str)
    """
    input_ext = input_ext.lower()
    output_fmt = output_fmt.lower()

    try:
        # IMAGE conversions (using Pillow)
        if input_ext in ["jpg", "jpeg", "png", "webp", "bmp", "gif", "tiff", "tif", "ico", "heic", "heif"] \
                and output_fmt in ["jpg", "jpeg", "png", "webp", "bmp", "gif", "tiff", "ico"]:
            return _convert_image(input_path, output_path, output_fmt)

        # SVG to image
        if input_ext == "svg" and output_fmt in ["jpg", "png"]:
            return _convert_svg_to_image(input_path, output_path, output_fmt)

        # Image to PDF
        if input_ext in ["jpg", "jpeg", "png", "webp", "bmp", "tiff", "tif"] and output_fmt == "pdf":
            return _convert_image_to_pdf(input_path, output_path)

        # GIF to MP4
        if input_ext == "gif" and output_fmt == "mp4":
            return _ffmpeg_convert(input_path, output_path, ["-movflags", "+faststart", "-pix_fmt", "yuv420p"])

        # AUDIO conversions (using ffmpeg)
        audio_formats = ["mp3", "wav", "ogg", "flac", "aac", "m4a", "opus", "wma"]
        if input_ext in audio_formats and output_fmt in audio_formats:
            return _convert_audio(input_path, output_path, output_fmt)

        # VIDEO conversions (using ffmpeg)
        video_formats = ["mp4", "avi", "mov", "mkv", "webm", "flv", "wmv", "3gp"]
        if input_ext in video_formats and output_fmt in video_formats:
            return _convert_video(input_path, output_path, output_fmt)

        # VIDEO to GIF
        if input_ext in video_formats and output_fmt == "gif":
            return _convert_video_to_gif(input_path, output_path)

        # VIDEO to AUDIO (extract audio)
        if input_ext in video_formats and output_fmt in audio_formats:
            return _ffmpeg_convert(input_path, output_path, ["-vn"])

        # DOCUMENT conversions (using LibreOffice)
        doc_formats = ["docx", "doc", "odt", "rtf", "pptx", "ppt", "xlsx", "xls"]
        if input_ext in doc_formats and output_fmt == "pdf":
            return _convert_with_libreoffice(input_path, output_path, "pdf")

        if input_ext in doc_formats and output_fmt in ["txt", "html", "docx", "csv"]:
            return _convert_with_libreoffice(input_path, output_path, output_fmt)

        # TXT to PDF / HTML
        if input_ext == "txt" and output_fmt == "pdf":
            return _convert_txt_to_pdf(input_path, output_path)

        if input_ext == "txt" and output_fmt == "html":
            return _convert_txt_to_html(input_path, output_path)

        if input_ext in ["html", "htm"] and output_fmt == "pdf":
            return _convert_html_to_pdf(input_path, output_path)

        if input_ext in ["html", "htm"] and output_fmt == "txt":
            return _convert_html_to_txt(input_path, output_path)

        # PDF to DOCX
        if input_ext == "pdf" and output_fmt == "docx":
            return _convert_pdf_to_docx(input_path, output_path)

        # PDF to TXT
        if input_ext == "pdf" and output_fmt == "txt":
            return _convert_pdf_to_txt(input_path, output_path)

        # PDF to image
        if input_ext == "pdf" and output_fmt in ["jpg", "png"]:
            return _convert_pdf_to_image(input_path, output_path, output_fmt)

        # CSV to XLSX
        if input_ext == "csv" and output_fmt == "xlsx":
            return _convert_csv_to_xlsx(input_path, output_path)

        # XLSX to CSV
        if input_ext in ["xlsx", "xls"] and output_fmt == "csv":
            return _convert_xlsx_to_csv(input_path, output_path)

        return False, f"Conversion from {input_ext} to {output_fmt} is not supported."

    except Exception as e:
        return False, str(e)


# ─── IMAGE CONVERTERS ─────────────────────────────────────────────────────────

def _convert_image(input_path, output_path, output_fmt):
    from PIL import Image
    fmt_map = {"jpg": "JPEG", "jpeg": "JPEG", "png": "PNG", "webp": "WEBP",
               "bmp": "BMP", "gif": "GIF", "tiff": "TIFF", "ico": "ICO"}
    pil_fmt = fmt_map.get(output_fmt, output_fmt.upper())
    img = Image.open(input_path)
    if pil_fmt == "JPEG" and img.mode in ("RGBA", "P", "LA"):
        img = img.convert("RGB")
    if pil_fmt == "ICO":
        img.save(output_path, format=pil_fmt, sizes=[(256, 256), (128, 128), (64, 64), (32, 32), (16, 16)])
    else:
        img.save(output_path, format=pil_fmt)
    return True, ""


def _convert_svg_to_image(input_path, output_path, output_fmt):
    try:
        import cairosvg
        if output_fmt == "png":
            cairosvg.svg2png(url=input_path, write_to=output_path)
        else:
            cairosvg.svg2png(url=input_path, write_to=output_path + ".tmp.png")
            _convert_image(output_path + ".tmp.png", output_path, output_fmt)
            os.remove(output_path + ".tmp.png")
        return True, ""
    except ImportError:
        return _ffmpeg_convert(input_path, output_path, [])


def _convert_image_to_pdf(input_path, output_path):
    from PIL import Image
    img = Image.open(input_path)
    if img.mode in ("RGBA", "P", "LA"):
        img = img.convert("RGB")
    img.save(output_path, "PDF", resolution=100.0)
    return True, ""


# ─── AUDIO CONVERTERS ─────────────────────────────────────────────────────────

def _convert_audio(input_path, output_path, output_fmt):
    extra = []
    if output_fmt == "mp3":
        extra = ["-codec:a", "libmp3lame", "-q:a", "2"]
    elif output_fmt == "ogg":
        extra = ["-codec:a", "libvorbis"]
    elif output_fmt == "flac":
        extra = ["-codec:a", "flac"]
    elif output_fmt == "aac":
        extra = ["-codec:a", "aac"]
    elif output_fmt == "m4a":
        extra = ["-codec:a", "aac"]
    elif output_fmt == "opus":
        extra = ["-codec:a", "libopus"]
    return _ffmpeg_convert(input_path, output_path, extra)


# ─── VIDEO CONVERTERS ─────────────────────────────────────────────────────────

def _convert_video(input_path, output_path, output_fmt):
    extra = []
    if output_fmt == "mp4":
        extra = ["-codec:v", "libx264", "-crf", "23", "-preset", "fast",
                 "-codec:a", "aac", "-movflags", "+faststart"]
    elif output_fmt == "webm":
        extra = ["-codec:v", "libvpx-vp9", "-crf", "30", "-b:v", "0", "-codec:a", "libopus"]
    elif output_fmt == "gif":
        return _convert_video_to_gif(input_path, output_path)
    return _ffmpeg_convert(input_path, output_path, extra)


def _convert_video_to_gif(input_path, output_path):
    # High quality GIF using palette
    palette_path = output_path + ".palette.png"
    cmd1 = ["ffmpeg", "-y", "-i", input_path,
             "-vf", "fps=10,scale=480:-1:flags=lanczos,palettegen",
             palette_path]
    cmd2 = ["ffmpeg", "-y", "-i", input_path, "-i", palette_path,
             "-filter_complex", "fps=10,scale=480:-1:flags=lanczos[x];[x][1:v]paletteuse",
             output_path]
    result1 = subprocess.run(cmd1, capture_output=True, text=True)
    if result1.returncode != 0:
        return False, result1.stderr[-500:]
    result2 = subprocess.run(cmd2, capture_output=True, text=True)
    if os.path.exists(palette_path):
        os.remove(palette_path)
    if result2.returncode != 0:
        return False, result2.stderr[-500:]
    return True, ""


# ─── DOCUMENT CONVERTERS ──────────────────────────────────────────────────────

def _convert_with_libreoffice(input_path, output_path, output_fmt):
    import shutil as sh
    lo = sh.which("libreoffice") or sh.which("soffice")
    if not lo:
        return False, "LibreOffice is not installed. Cannot convert documents."
    out_dir = str(Path(output_path).parent)
    fmt_map = {"pdf": "pdf", "txt": "txt", "html": "html", "docx": "docx", "csv": "csv"}
    lo_fmt = fmt_map.get(output_fmt, output_fmt)
    cmd = [lo, "--headless", "--convert-to", lo_fmt, "--outdir", out_dir, input_path]
    result = subprocess.run(cmd, capture_output=True, text=True, timeout=120)
    if result.returncode != 0:
        return False, result.stderr[-500:]
    # LibreOffice names the file based on input name, rename to expected output
    stem = Path(input_path).stem
    lo_output = Path(out_dir) / f"{stem}.{lo_fmt}"
    if lo_output.exists() and str(lo_output) != output_path:
        lo_output.rename(output_path)
    return True, ""


def _convert_txt_to_pdf(input_path, output_path):
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.pdfgen import canvas
        from reportlab.lib.units import cm
        with open(input_path, "r", encoding="utf-8", errors="replace") as f:
            lines = f.readlines()
        c = canvas.Canvas(output_path, pagesize=A4)
        width, height = A4
        margin = 2 * cm
        y = height - margin
        c.setFont("Helvetica", 11)
        for line in lines:
            line = line.rstrip("\n")
            if y < margin:
                c.showPage()
                c.setFont("Helvetica", 11)
                y = height - margin
            c.drawString(margin, y, line[:100])
            y -= 14
        c.save()
        return True, ""
    except ImportError:
        return _convert_with_libreoffice(input_path, output_path, "pdf")


def _convert_txt_to_html(input_path, output_path):
    with open(input_path, "r", encoding="utf-8", errors="replace") as f:
        content = f.read()
    html = f"""<!DOCTYPE html>
<html><head><meta charset="utf-8"><title>Converted</title>
<style>body{{font-family:monospace;white-space:pre-wrap;padding:20px}}</style>
</head><body>{content}</body></html>"""
    with open(output_path, "w", encoding="utf-8") as f:
        f.write(html)
    return True, ""


def _convert_html_to_pdf(input_path, output_path):
    try:
        import weasyprint
        weasyprint.HTML(filename=input_path).write_pdf(output_path)
        return True, ""
    except ImportError:
        return _convert_with_libreoffice(input_path, output_path, "pdf")


def _convert_html_to_txt(input_path, output_path):
    try:
        from bs4 import BeautifulSoup
        with open(input_path, "r", encoding="utf-8", errors="replace") as f:
            soup = BeautifulSoup(f.read(), "html.parser")
        with open(output_path, "w", encoding="utf-8") as f:
            f.write(soup.get_text())
        return True, ""
    except ImportError:
        return False, "beautifulsoup4 not installed."


def _convert_pdf_to_docx(input_path, output_path):
    try:
        from pdf2docx import Converter
        cv = Converter(input_path)
        cv.convert(output_path, start=0, end=None)
        cv.close()
        return True, ""
    except ImportError:
        return False, "pdf2docx not installed. Run: pip install pdf2docx"


def _convert_pdf_to_txt(input_path, output_path):
    try:
        import pdfplumber
        text = ""
        with pdfplumber.open(input_path) as pdf:
            for page in pdf.pages:
                text += (page.extract_text() or "") + "\n\n"
        with open(output_path, "w", encoding="utf-8") as f:
            f.write(text)
        return True, ""
    except ImportError:
        try:
            import PyPDF2
            with open(input_path, "rb") as f:
                reader = PyPDF2.PdfReader(f)
                text = "\n".join(page.extract_text() or "" for page in reader.pages)
            with open(output_path, "w", encoding="utf-8") as f:
                f.write(text)
            return True, ""
        except ImportError:
            return False, "pdfplumber or PyPDF2 not installed."


def _convert_pdf_to_image(input_path, output_path, output_fmt):
    try:
        from pdf2image import convert_from_path
        images = convert_from_path(input_path, dpi=150, first_page=1, last_page=1)
        if images:
            fmt = "JPEG" if output_fmt == "jpg" else "PNG"
            if fmt == "JPEG":
                images[0] = images[0].convert("RGB")
            images[0].save(output_path, fmt)
        return True, ""
    except ImportError:
        return False, "pdf2image not installed. Also needs poppler-utils system package."


def _convert_csv_to_xlsx(input_path, output_path):
    import csv
    try:
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        with open(input_path, "r", encoding="utf-8", errors="replace") as f:
            reader = csv.reader(f)
            for row in reader:
                ws.append(row)
        wb.save(output_path)
        return True, ""
    except ImportError:
        return False, "openpyxl not installed."


def _convert_xlsx_to_csv(input_path, output_path):
    try:
        import openpyxl
        import csv
        wb = openpyxl.load_workbook(input_path, read_only=True, data_only=True)
        ws = wb.active
        with open(output_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            for row in ws.iter_rows(values_only=True):
                writer.writerow(row)
        return True, ""
    except ImportError:
        return False, "openpyxl not installed."


# ─── FFMPEG HELPER ────────────────────────────────────────────────────────────

def _ffmpeg_convert(input_path, output_path, extra_args):
    cmd = ["ffmpeg", "-y", "-i", input_path] + extra_args + [output_path]
    result = subprocess.run(cmd, capture_output=True, text=True, timeout=300)
    if result.returncode != 0:
        return False, result.stderr[-500:]
    return True, ""
